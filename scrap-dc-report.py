import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font

st.title("Excel Formatter App")

# File uploader
uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    # Read Excel
    df_raw = pd.read_excel(uploaded_file, header=None)

    # Required columns
    required_cols = ["Wcgroup Name", "Wc Name", "State", "Karat Code", "Weight", "Pg Weight"]

    # Find header row dynamically
    header_row = None
    for i, row in df_raw.iterrows():
        if all(col in row.values for col in required_cols):
            header_row = i
            break

    if header_row is not None:
        # Read again with correct header
        df = pd.read_excel(uploaded_file, header=header_row)

        # Keep only required columns in same order
        df = df[[col for col in required_cols if col in df.columns]]

        st.write("✅ Formatted Table Preview", df.head())

        # Save to BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Formatted")
            workbook = writer.book
            worksheet = writer.sheets["Formatted"]

            # Column widths
            col_widths = {"A": 24, "B": 24}
            for col, width in col_widths.items():
                worksheet.column_dimensions[col].width = width

            # Borders
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                           min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(
                        vertical="center",
                        horizontal="center",
                        wrap_text=True
                    )

            # Highlight Top Header Row
            worksheet.insert_rows(1)
            worksheet["A1"] = "Stock Status :- Scrap _ Date : 03-09-2025  Time : 10:10"
            worksheet.merge_cells(start_row=1, start_column=1,
                                  end_row=1, end_column=worksheet.max_column)

            header_cell = worksheet["A1"]
            header_cell.font = Font(bold=True, size=12, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.row_dimensions[1].height = 25

            # Bottom total row formatting
            last_row = worksheet.max_row
            for cell in worksheet[last_row]:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            worksheet.row_dimensions[last_row].height = 35

        # Download button
        st.download_button(
            label="📥 Download Formatted Excel",
            data=output.getvalue(),
            file_name="Formatted_Stock.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("❌ Could not find required columns in uploaded file.")
