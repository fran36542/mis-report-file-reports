import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from io import BytesIO

st.title("Excel Formatter - Delay Order Report")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx", "xls"])

if uploaded_file:
    # 1️⃣ Read Excel safely, preserve all data as string
    try:
        df = pd.read_excel(
            uploaded_file, header=None, dtype=str, keep_default_na=False, na_values=[]
        )
    except Exception as e:
        st.error(f"Error reading Excel: {e}")
        st.stop()

    # 2️⃣ Column keywords mapping
    column_keywords = {
        "Order No": ["order no", "orderno", "order #"],
        "Variant Name": ["variant", "variant name"],
        "Group No": ["group no", "group"],
        "Batch No": ["batch no", "batch"],
        "Actual Metal Wt": ["metal wt", "actual metal"],
        "Dept Name": ["dept", "department"],
        "Jobworker": ["jobworker", "worker"],
        "Req Purity": ["req purity", "purity"],
        "Shipment Date": ["shipment date", "ship date"],
        "Order Date": ["order date", "date"],
        "Shipment Days": ["shipment days", "ship days"],
        "Order Days": ["order days", "days"]
    }

    # 3️⃣ Detect header row dynamically
    header_row_index = None
    for i, row in df.iterrows():
        row_lower = [str(x).strip().lower() for x in row]
        matches = sum(
            any(kw in str(x).strip().lower() for kw in keywords) 
            for x in row for key, keywords in column_keywords.items()
        )
        if matches >= len(column_keywords) // 2:
            header_row_index = i
            break
    if header_row_index is None:
        st.warning("Header row not found. Using first row as header.")
        header_row_index = 0

    # 4️⃣ Map columns dynamically
    header_row = df.iloc[header_row_index].astype(str)
    new_columns = []
    for col in header_row:
        col_lower = str(col).strip().lower()
        mapped = None
        for key, keywords in column_keywords.items():
            if any(kw in col_lower for kw in keywords):
                mapped = key
                break
        new_columns.append(mapped if mapped else str(col))
    df = df.iloc[header_row_index + 1:]  # data below header
    df.columns = new_columns

    # 5️⃣ Ensure all required columns exist
    for col in column_keywords.keys():
        if col not in df.columns:
            df[col] = ""
    df = df[list(column_keywords.keys())]
    df.reset_index(drop=True, inplace=True)

    # 6️⃣ Remove bottom "Grand Total" rows
    df = df[~df.apply(lambda row: row.astype(str).str.contains("grand total", case=False).any(), axis=1)]

    # 7️⃣ Filter Order Days: remove blanks and (1.00) to (9.00)
    unwanted_values = ["", "(1.00)", "(2.00)", "(3.00)", "(4.00)", "(5.00)", "(6.00)", "(7.00)", "(8.00)", "(9.00)"]
    df = df[~df["Order Days"].astype(str).isin(unwanted_values)]

    # 8️⃣ Sort Order Days descending (keep existing string formatting like (10.00))
    # Convert Order Days to float for sorting but keep original string
    df["Order Days Sort"] = df["Order Days"].str.replace("(", "").str.replace(")", "").astype(float)
    df.sort_values(by="Order Days Sort", ascending=False, inplace=True)
    df.drop(columns=["Order Days Sort"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    # 9️⃣ Format Shipment Date and Order Date
    for col in ["Shipment Date", "Order Date"]:
        df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d-%m-%Y')

    # 🔟 Save to Excel with formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=2)
        ws = writer.sheets['Sheet1']

        # Top headers
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws["A1"] = "DELAY ORDER REPORT"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(bold=True, size=14)

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
        ws["A2"] = "Status :- OPEN"
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"].font = Font(bold=True)

        # Header formatting
        header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = 15
            cell.border = thin_border

        # Data formatting with borders
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                cell.border = thin_border

        # Header row height
        ws.row_dimensions[3].height = 25

    output.seek(0)
    st.download_button(
        label="Download Formatted Excel",
        data=output,
        file_name="Formatted_Delay_Order_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )