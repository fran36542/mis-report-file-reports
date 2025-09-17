import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
import tempfile
import os

st.set_page_config(page_title="Excel Formatter", layout="wide")

# Custom CSS for styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .instruction-box {
        background-color: #f0f7ff;
        padding: 20px;
        border-radius: 10px;
        border-left: 5px solid #1f77b4;
        margin-bottom: 20px;
    }
    .success-msg {
        color: #28a745;
        font-weight: bold;
    }
    .stButton>button {
        background-color: #1f77b4;
        color: white;
        border: none;
        padding: 10px 20px;
        border-radius: 5px;
        font-weight: bold;
    }
    .stButton>button:hover {
        background-color: #1668a5;
    }
    .download-btn {
        background-color: #28a745 !important;
    }
    .download-btn:hover {
        background-color: #218838 !important;
    }
    .preview-container {
        border: 1px solid #ddd;
        border-radius: 5px;
        padding: 10px;
        margin-top: 20px;
    }
</style>
""", unsafe_allow_html=True)

# App title and description
st.markdown('<h1 class="main-header">Excel Formatting Tool</h1>', unsafe_allow_html=True)

st.markdown("""
<div class="instruction-box">
    <h3>📋 How to Use</h3>
    <ol>
        <li>Upload your Excel file using the uploader below</li>
        <li>The application will preserve your original formatting</li>
        <li>We'll add borders to all cells (including empty ones) for a complete look</li>
        <li>Download the formatted Excel file with the button provided</li>
    </ol>
    <p><strong>Note:</strong> We maintain your original content and layout, only adding borders for better visualization.</p>
</div>
""", unsafe_allow_html=True)

# File upload section
uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])

def add_borders_to_worksheet(ws):
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            # Apply border to every cell
            cell.border = thin_border
            
            # Preserve existing horizontal alignment or set to left
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal='left')

def process_excel(file):
    # Read the Excel file
    wb = openpyxl.load_workbook(file)
    
    # Process each worksheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        add_borders_to_worksheet(ws)
    
    # Save to a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    wb.close()
    
    return temp_file.name

if uploaded_file is not None:
    try:
        # Process the file
        with st.spinner('Processing your Excel file...'):
            processed_file_path = process_excel(uploaded_file)
        
        # Success message
        st.success('✅ File processed successfully! Borders have been added to all cells while preserving your original formatting.')
        
        # Display a preview of the Excel file
        st.subheader("Preview of Your Excel File")
        
        # Read and display the first sheet as a dataframe
        df = pd.read_excel(processed_file_path)
        
        # Create a styled dataframe with borders for the preview
        styled_df = df.style.set_properties(**{
            'border': '1px solid black',
            'text-align': 'left',
            'padding': '5px'
        }).set_table_styles([{
            'selector': 'th',
            'props': [('border', '1px solid black'), ('background-color', '#f0f7ff')]
        }])
        
        # Display the styled dataframe
        st.dataframe(styled_df, use_container_width=True, height=400)
        
        # Download button
        with open(processed_file_path, 'rb') as f:
            excel_data = f.read()
        
        st.download_button(
            label="📥 Download Formatted Excel File",
            data=excel_data,
            file_name="formatted_excel.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download-excel",
            help="Click to download the formatted Excel file with borders on all cells"
        )
        
        # Clean up temporary file
        os.unlink(processed_file_path)
        
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
else:
    # Display sample data when no file is uploaded
    st.info("👆 Please upload an Excel file to get started. Here's a sample of what the formatting will look like:")
    
    # Create sample data similar to the user's example with all arrays of the same length
    sample_data = {
        'Column1': ['', 'SR.FAC', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        'Column2': ['', 'Stock MLS Statement (Without Value)~1 SR', 'Item Type : METAL', 
                  'Item Group : GOLD', 'From Date : 08/09/2025', 'To Date : 08/09/2025', 
                  'Company Name : SR', '', 'GOLD IN-OUT SUMMARY', '', 'PARTICULARS', 
                  'SR.FAC', 'ADD', 'ACCEPTANCE (DEPARTMENT)', 'CONVERT TO AVA METAL PRINCESS', 
                  'GOODS RECEPT NOTE', 'KARAT CONTASSION', 'LOSS RECOVERY', 'DPS -', 
                  'RETURN SCRAP METAL', 'STOCK RECONCILATION', 'TOTAL ADD'],
        'Column3': ['', '', '', '', '', '', '', '', '', '', '', '', '', 
                   98728.427, 0.158, 12012.558, 12.542, 43.889, 303978.568, 
                   0.152, 4.269, 414920.552]
    }
    
    sample_df = pd.DataFrame(sample_data)
    
    # Style the sample dataframe to show borders
    styled_sample = sample_df.style.set_properties(**{
        'border': '1px solid black',
        'text-align': 'left',
        'padding': '5px'
    }).set_table_styles([{
        'selector': 'th',
        'props': [('border', '1px solid black'), ('background-color', '#f0f7ff')]
    }])
    
    st.dataframe(styled_sample, use_container_width=True, hide_index=True)

# Footer
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #666;'>Excel Formatting Tool • Upload, Format, and Download Your Excel Files</div>", 
    unsafe_allow_html=True
)