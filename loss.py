import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from io import BytesIO
import re

# Set page configuration
st.set_page_config(
    page_title="Excel Loss Report Processor",
    page_icon="📊",
    layout="centered"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        text-align: center;
        color: #2c3e50;
        margin-bottom: 30px;
    }
    .stButton>button {
        width: 100%;
        background-color: #2ecc71;
        color: white;
    }
    .uploadedFile {
        display: none;
    }
</style>
""", unsafe_allow_html=True)

required_columns = [
    "Wc Name",
    "Issue Quantity Pg",
    "Process Quantity Pg",
    "Unutilized Quantity Pg",
    "Unutilized Quantity Sample Pg",
    "Unutilized Quantity Scrap Pg",
    "Loss Quantity Pg",
    "Gain Pg",
    "FINAL LOSS",
    "LOSS %"
]

# Custom column widths
col_widths = {
    "Wc Name": 30,
    "Issue Quantity Pg": 12,
    "Process Quantity Pg": 14,
    "Unutilized Quantity Pg": 11.5,
    "Unutilized Quantity Sample Pg": 17.3,
    "Unutilized Quantity Scrap Pg": 17.3,
    "Loss Quantity Pg": 11.5,
    "Gain Pg": 11.5,
    "FINAL LOSS": 11.5,
    "LOSS %": 11.5
}

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def extract_dates_from_excel(data):
    """Extract From Date and To Date from uploaded Excel file"""
    from_date, to_date = None, None

    # scan first 10 rows (header/top area)
    for row in data[:10]:
        row_str = [str(cell).strip() if cell else "" for cell in row]

        for i, cell in enumerate(row_str):
            low = cell.lower()

            # ---- detect "from date" ----
            if "from date" in low:
                if ":" in cell:
                    part = cell.split(":")[-1].strip()
                    if part:
                        from_date = part
                elif i + 1 < len(row_str) and row_str[i + 1]:
                    from_date = row_str[i + 1]

            # ---- detect "to date" ----
            if "to date" in low:
                if ":" in cell:
                    part = cell.split(":")[-1].strip()
                    if part:
                        to_date = part
                elif i + 1 < len(row_str) and row_str[i + 1]:
                    to_date = row_str[i + 1]

    return from_date, to_date

def process_excel(file_stream):
    wb_old = openpyxl.load_workbook(file_stream, data_only=False)
    ws_old = wb_old.active
    data = list(ws_old.values)

    # --- Extract date from uploaded file (else fallback) ---
    from_date, to_date = extract_dates_from_excel(data)
    if not from_date:
        from_date = datetime.now().strftime("%d/%m/%Y")
    if not to_date:
        to_date = datetime.now().strftime("%d/%m/%Y")

    wb_new = openpyxl.Workbook()
    ws = wb_new.active
    ws.title = "Formatted Loss Report"

    # --- Title and Date Rows ---
    titles = [
        "Monthly Loss Report Summary",
        f"From Date :- {from_date}",
        f"To Date :- {to_date}"
    ]

    for i, text in enumerate(titles, start=1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(required_columns))
        cell = ws.cell(i, 1, text)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if i == 1:
            cell.font = Font(bold=True, size=16, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0000FF")
            ws.row_dimensions[i].height = 25  # Title height

    # --- Header row ---
    header_row = 4
    for col_num, col_name in enumerate(required_columns, 1):
        cell = ws.cell(header_row, col_num, col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="ADD8E6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = col_widths.get(col_name, 12)

    ws.row_dimensions[header_row].height = 37.5  # Header height

    # --- Find header row in uploaded file ---
    header_row_idx = -1
    for i, row in enumerate(data):
        if any(str(cell).strip().lower() == 'wc name' for cell in row if cell):
            header_row_idx = i
            break
    if header_row_idx == -1:
        raise Exception("Could not find header row with 'Wc Name'")

    headers_old = [str(c).strip() if c else "" for c in data[header_row_idx]]
    col_index_map = {}
    for col in required_columns[:-2]:
        found_idx = None
        for idx, h in enumerate(headers_old):
            if h and col.lower().replace(" ", "") in str(h).lower().replace(" ", ""):
                found_idx = idx
                break
        col_index_map[col] = found_idx

    # --- Extract rows (keep TOTAL row from user file) ---
    data_rows = data[header_row_idx + 1:]
    data_rows = [row for row in data_rows if any(row)]

    start_row = 5

    # --- Write rows ---
    for row_idx, old_row in enumerate(data_rows, start=start_row):
        # detect TOTAL row
        is_total_row = False
        try:
            cell0 = old_row[0]
            if cell0 is not None and isinstance(cell0, str) and cell0.strip().lower().startswith("total"):
                is_total_row = True
        except Exception:
            is_total_row = False

        for col_idx, col_name in enumerate(required_columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if is_total_row:
                # Copy values, highlight yellow
                old_idx = col_index_map.get(col_name)
                if old_idx is not None and old_idx < len(old_row):
                    cell.value = old_row[old_idx]
                cell.fill = PatternFill("solid", fgColor="FFFF00")
                cell.font = Font(bold=True)
            else:
                # Normal data rows with formulas
                if col_name == "FINAL LOSS":
                    cell.value = f"=G{row_idx}+H{row_idx}"
                    cell.number_format = '0.0000'
                    cell.fill = PatternFill("solid", fgColor="FFFF99")
                elif col_name == "LOSS %":
                    cell.value = f"=I{row_idx}/(C{row_idx}+D{row_idx}+E{row_idx}+F{row_idx})"
                    cell.number_format = '0.00%'
                    cell.fill = PatternFill("solid", fgColor="FFFF99")
                else:
                    old_idx = col_index_map.get(col_name)
                    if old_idx is not None and old_idx < len(old_row):
                        val = old_row[old_idx]
                        if isinstance(val, (int, float)):
                            cell.value = round(val, 4)
                            cell.number_format = '0.0000'
                        else:
                            cell.value = val
            cell.border = thin_border

    # ---- AFTER writing rows: ensure TOTAL row height = 23 ----
    for r in range(start_row, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        if a_val and isinstance(a_val, str) and a_val.strip().lower().startswith("total"):
            ws.row_dimensions[r].height = 23
            try:
                ws.row_dimensions[r].customHeight = True
            except Exception:
                pass
            for c in range(1, len(required_columns) + 1):
                ccell = ws.cell(r, c)
                ccell.fill = PatternFill("solid", fgColor="FFFF00")
                ccell.font = Font(bold=True)
                ccell.border = thin_border

    # --- Save output ---
    output = BytesIO()
    wb_new.save(output)
    output.seek(0)
    return output

def main():
    st.markdown("<h1 class='main-header'>Excel Loss Report Processor</h1>", unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader("Choose Excel File", type=["xlsx"], accept_multiple_files=False)
    
    if uploaded_file is not None:
        if st.button("Process File"):
            with st.spinner("Processing file... Please wait."):
                try:
                    processed_file = process_excel(uploaded_file)
                    
                    date_str = datetime.now().strftime("%Y%m%d")
                    filename = f"loss_report_{date_str}.xlsx"
                    
                    st.success("File processed successfully!")
                    
                    # Download button
                    st.download_button(
                        label="Download Processed File",
                        data=processed_file,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Error processing file: {str(e)}")

if __name__ == "__main__":
    main()